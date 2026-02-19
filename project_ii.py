from opencv.fr import FR
from opencv.fr.search.schemas import SearchRequest, SearchMode

import cv2
import tempfile
import os
from pathlib import Path
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pyttsx3

# ================= CONFIG =================

engine = pyttsx3.init()

BACKEND_URL = "https://us.opencv.fr"
DEVELOPER_KEY = "elaBl7xMjVlNGI2ZmUtZTA1YS00MWRiLWE3N2QtMjdiMDhhY2M5NTc4"

API_COOLDOWN = 2.5
MIN_EXIT_GAP = 20 * 60
REENTRY_BLOCK = 21 * 3600

excel_file = "attendence_log.xlsx"

# ===========================================

sdk = FR(BACKEND_URL, DEVELOPER_KEY)

person_memory = {}

face_detector = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

# ================= SPEECH ===================

def speak_once(text):
    engine.say(text)
    engine.runAndWait()

# ================= EXCEL ====================

def init_excel():
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["date", "entry_time", "exit_time", "person_id", "name", "duration"])
        wb.save(excel_file)

def format_duration(seconds):
    hrs = int(seconds // 3600)
    mins = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    return f"{hrs:02}:{mins:02}:{secs:02}"

# ✅ WRITE ENTRY IMMEDIATELY
def log_entry(person_id, name, entry_time):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        date_str = datetime.now().strftime("%d-%m-%Y")
        entry_str = datetime.fromtimestamp(entry_time).strftime("%H:%M:%S")

        ws.append([date_str, entry_str, None, person_id, name, None])
        wb.save(excel_file)

        print(f"✅ Entry Logged → {name}")

    except PermissionError:
        speak_once("Close Excel file")

# ✅ UPDATE EXIT IN SAME ROW
def log_exit_update(person_id, exit_time):

    wb = load_workbook(excel_file)
    ws = wb.active

    exit_str = datetime.fromtimestamp(exit_time).strftime("%H:%M:%S")

    for row in reversed(list(ws.iter_rows(min_row=2))):

        row_person_id = row[3].value
        row_exit = row[2].value

        if row_person_id == person_id and row_exit is None:

            entry_time = datetime.strptime(
                f"{row[0].value} {row[1].value}",
                "%d-%m-%Y %H:%M:%S"
            ).timestamp()

            duration_seconds = exit_time - entry_time
            duration_str = format_duration(duration_seconds)

            row[2].value = exit_str
            row[5].value = duration_str

            wb.save(excel_file)

            print("✅ Exit Updated")
            return

# ✅ RESTORE STATE AFTER RESTART
def restore_memory_excel():
    if not os.path.exists(excel_file):
        return

    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):

        if not row or len(row) < 6:
            continue

        date, entry_time, exit_time, person_id, name, duration = row

        if not person_id:
            continue

        if person_id not in person_memory:
            person_memory[person_id] = {
                "name": name,
                "last_entry_time": None,
                "last_exit_time": None
            }

        if entry_time:
            entry_ts = datetime.strptime(
                f"{date} {entry_time}", "%d-%m-%Y %H:%M:%S"
            ).timestamp()
            person_memory[person_id]["last_entry_time"] = entry_ts

        if exit_time:
            exit_ts = datetime.strptime(
                f"{date} {exit_time}", "%d-%m-%Y %H:%M:%S"
            ).timestamp()
            person_memory[person_id]["last_exit_time"] = exit_ts

# ================= VISION ===================

def detect_face(frame):
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_detector.detectMultiScale(gray, 1.3, 5)
    return len(faces) > 0

# ================= API ======================

def search_frame(frame):
    temp_path = None

    try:
        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
            temp_path = tmp.name

        cv2.imwrite(temp_path, frame)

        search_request = SearchRequest(
            [Path(temp_path)],
            min_score=0.6,
            collection_id=None,
            search_mode=SearchMode.FAST
        )

        return sdk.search.search(search_request)

    except Exception as e:
        print("🔥 API Error:", e)

    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)

    return None

# ================= LOGIC ====================

def update_logic(results):

    current_time = time.time()
    now_str = datetime.now().strftime("%H:%M:%S")

    if not results:
        print(f"[{now_str}] ❌ Person Not Registered")
        speak_once("Person not registered")
        return "unregistered"

    for match in results:

        person = match.person
        person_id = person.id
        name = person.name

        if person_id not in person_memory:

            person_memory[person_id] = {
                "name": name,
                "last_entry_time": current_time,
                "last_exit_time": None
            }

            print(f"[{now_str}] 🏢 ENTRY → {name}")
            log_entry(person_id, name, current_time)
            speak_once("Good morning, you are checked in")

            return "entry"

        memory = person_memory[person_id]
        last_entry = memory["last_entry_time"]
        last_exit = memory["last_exit_time"]

        # ===== EXIT =====
        if last_entry and last_exit is None:

            time_inside = current_time - last_entry

            if time_inside >= MIN_EXIT_GAP:

                memory["last_exit_time"] = current_time

                print(f"[{now_str}] 🚪 EXIT → {name}")
                log_exit_update(person_id, current_time)
                speak_once("Checkout successful")

                return "exit"

            else:
                remaining = int((MIN_EXIT_GAP - time_inside) // 60)
                print(f"[{now_str}] ⛔ EXIT BLOCKED → {name}")
                speak_once("You cannot exit yet")
                return "blocked_exit"

        # ===== REENTRY =====
        elif last_exit:

            time_since_exit = current_time - last_exit

            if time_since_exit >= REENTRY_BLOCK:

                memory["last_entry_time"] = current_time
                memory["last_exit_time"] = None

                print(f"[{now_str}] 🏢 RE-ENTRY → {name}")
                log_entry(person_id, name, current_time)
                speak_once("Welcome back")

                return "reentry"

            else:
                print(f"[{now_str}] ⛔ RE-ENTRY BLOCKED → {name}")
                speak_once("Re-entry blocked")
                return "blocked_reentry"

# ================= CAMERA ===================

def run_camera():

    cap = cv2.VideoCapture(0)

    init_excel()
    restore_memory_excel()

    if not cap.isOpened():
        print("❌ Cannot open camera")
        return

    print("🚀 Live Recognition Started | Press 'q' to quit")

    last_api_call = 0
    status_text = ""

    while True:

        ret, frame = cap.read()
        if not ret:
            break

        frame = cv2.flip(frame, 1)
        frame = cv2.resize(frame, (640, 480))

        current_time = time.time()

        if detect_face(frame):

            if current_time - last_api_call > API_COOLDOWN:

                results = search_frame(frame)
                state = update_logic(results)

                status_text = state
                last_api_call = current_time

        cv2.imshow("Office Entry/Exit System", frame)

        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()

# ============================================

if __name__ == "__main__":
    run_camera()
